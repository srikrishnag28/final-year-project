from django.shortcuts import render, HttpResponse, redirect
from .models import Criteria_3_1_1, Criteria_3_2_2
from .forms import *
from datetime import datetime
from django.db import transaction
import openpyxl
from openpyxl import Workbook
from io import BytesIO
from django.contrib import messages
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from user.models import UserCriteriaLink
from django.contrib.auth.decorators import login_required

# Create your views here.


@login_required
def criteria_3_1_1(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='3_1_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="3_1_1")
    context = {
        'has_access': has_access,
        'users': users,
        '2021_records': Criteria_3_1_1.objects.filter(year = '2021'),
        '2022_records': Criteria_3_1_1.objects.filter(year = '2022'),
        '2023_records': Criteria_3_1_1.objects.filter(year = '2023'),
        '2024_records': Criteria_3_1_1.objects.filter(year = '2024'),
        '2025_records': Criteria_3_1_1.objects.filter(year = '2025')
    }
    return render(request, 'table/criteria_3_1_1.html', context=context)


@login_required
def criteria_3_1_1_form(request):
    if request.method == 'POST':
        form = CriteriaForm_3_1_1(request.POST)
        if form.is_valid():
            form.save() 
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_3_1_1')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_3_1_1()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_3_1_1.html', context=context)


@login_required
def upload_excel_3_1_1(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    year = row[0]  # Ensure this matches the year choices
                    project_name = row[1]
                    principal_investigator = row[2]
                    department = row[3]
                    year_of_award = row[4]
                    amount_sanctioned = row[5]
                    duration = row[6]
                    funding_agency = row[7]
                    grant_type = row[8]  # Ensure this is either 'Government' or 'Non-Government'

                    Criteria_3_1_1.objects.create(
                        year=year,
                        project_name=project_name,
                        principal_investigator=principal_investigator,
                        department=department,
                        year_of_award=year_of_award,
                        amount_sanctioned=amount_sanctioned,
                        duration=duration,
                        funding_agency=funding_agency,
                        grant_type=grant_type,
                    )
            messages.success(request, 'Excel File added successfully')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
        return redirect('criteria_3_1_1')
    messages.error(request, 'Upload a valid Excel file')
    return redirect('criteria_3_1_1')


@login_required
def export_excel_3_1_1(request):
    data = Criteria_3_1_1.objects.all().order_by('year')

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Criteria 3.1.1"

    headers = [
        'Name of the research project/ endowment',
        'Name of the Principal Investigator/Co-investigator',
        'Department of Principal Investigator',
        'Year of Award',
        'Amount Sanctioned',
        'Duration of the project',
        'Name of the Funding Agency',
        'Type (Government/non-Government)'
    ]

    column_width = 21.89
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = column_width

    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='top')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    bold_font = Font(bold=True)

    yellow_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

    title_cell = worksheet.cell(row=1, column=1, value='3.1.1 Grants received from Government and non-governmental'
                                                       ' agencies for research projects / endowments in the '
                                                       'institution during the last five years (INR in Lakhs) ')

    start_row = 3

    for year in Criteria_3_1_1.YEAR_CHOICES:
        year_value, year_label = year

        worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
        year_cell = worksheet.cell(row=start_row, column=1, value=f"Year: {year_label}")
        year_cell.alignment = center_alignment
        year_cell.font = bold_font
        year_cell.fill = yellow_fill

        for row in worksheet.iter_rows(min_row=start_row, max_row=start_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        start_row += 1

        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col_num, value=header)
            cell.alignment = wrap_alignment
            cell.border = thin_border
            cell.font = bold_font
        start_row += 1

        year_data = data.filter(year=year_value)
        for entry in year_data:
            cells = [
                worksheet.cell(row=start_row, column=1, value=entry.project_name),
                worksheet.cell(row=start_row, column=2, value=entry.principal_investigator),
                worksheet.cell(row=start_row, column=3, value=entry.department),
                worksheet.cell(row=start_row, column=4, value=entry.year_of_award),
                worksheet.cell(row=start_row, column=5, value=entry.amount_sanctioned),
                worksheet.cell(row=start_row, column=6, value=entry.duration),
                worksheet.cell(row=start_row, column=7, value=entry.funding_agency),
                worksheet.cell(row=start_row, column=8, value=entry.grant_type),
            ]

            for cell in cells:
                cell.alignment = wrap_alignment
                cell.border = thin_border
                cell.font = bold_font
            start_row += 1

        start_row += 1

    virtual_workbook = BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response = HttpResponse(
        virtual_workbook,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=3.1.1_criteria_data.xlsx'
    return response


@login_required
def criteria_3_2_1_form(request):
    # Get the existing instance (if any)
    instance = Criteria_3_2_1.objects.first()

    if request.method == 'POST':
        form = CriteriaForm_3_2_1(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('criteria_3_2_1')
        else:
            print(form.errors)
    else:
        form = CriteriaForm_3_2_1(instance=instance)

    return render(request, 'form/criteria_3_2_1.html', {'form': form})


@login_required
def criteria_3_2_1(request):
    criteria_data = Criteria_3_2_1.objects.first()
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='3_2_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="3_2_1")
    context = {
        'criteria_data': criteria_data,
        'has_access': has_access,
        'users': users,
    }
    if not criteria_data:
        context['no_data'] = True
    return render(request, 'table/criteria_3_2_1.html', context)


@login_required
def criteria_3_2_1_delete(request):
    Criteria_3_2_1.objects.all().delete()
    return redirect('criteria_3_2_1')


@login_required
def criteria_3_2_2(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='3_2_2').exists()
    users = UserCriteriaLink.objects.filter(criteria="3_2_2")
    context = {
        'has_access': has_access,
        'users': users,
        '2021_records': Criteria_3_2_2.objects.filter(year = '2021'),
        '2022_records': Criteria_3_2_2.objects.filter(year = '2022'),
        '2023_records': Criteria_3_2_2.objects.filter(year = '2023'),
        '2024_records': Criteria_3_2_2.objects.filter(year = '2024'),
        '2025_records': Criteria_3_2_2.objects.filter(year = '2025')
    }
    return render(request, 'table/criteria_3_2_2.html', context=context)


@login_required
def criteria_3_2_2_form(request):
    if request.method == 'POST':
        form = CriteriaForm_3_2_2(request.POST)
        if form.is_valid():
            form.save() 
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_3_2_2')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_3_2_2()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_3_2_2.html', context=context)


@login_required
def upload_excel_3_2_2(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    year = row[0]  # Ensure this matches the year choices
                    event_name = row[1]
                    num_participants = row[2]
                    date_from = row[3]  # Ensure this is a valid date
                    date_to = row[4]  # Ensure this is a valid date
                    activity_report_link = row[5]  # Ensure this is a valid URL

                    Criteria_3_2_2.objects.create(
                        year=year,
                        event_name=event_name,
                        num_participants=num_participants,
                        date_from=date_from,
                        date_to=date_to,
                        activity_report_link=activity_report_link,
                    )
            messages.success(request, 'Excel File added successfully')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
        return redirect('criteria_3_2_2')
    messages.error(request, 'Upload a valid Excel file')
    return redirect('criteria_3_2_2')


@login_required
def export_excel_3_2_2(request):
    data = Criteria_3_2_2.objects.all().order_by('year')

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Criteria 3.2.2"

    headers = [
        'Name of the workshop/ seminar/ conference',
        'Number of Participants',
        'Date (From – To)',
        'Link to the Activity report on the website'
    ]

    column_width = 21.89
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = column_width

    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='top')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    bold_font = Font(bold=True)

    yellow_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

    title_cell = worksheet.cell(row=1, column=1,
                                value='3.2.2 Number of workshops/seminars/conferences including programs conducted on'
                                      ' Research Methodology, Intellectual Property Rights (IPR) and entrepreneurship '
                                      'during the last five ')

    start_row = 3

    for year in Criteria_3_2_2.YEAR_CHOICES:
        year_value, year_label = year

        worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
        year_cell = worksheet.cell(row=start_row, column=1, value=f"Year: {year_label}")
        year_cell.alignment = center_alignment
        year_cell.font = bold_font
        year_cell.fill = yellow_fill

        for row in worksheet.iter_rows(min_row=start_row, max_row=start_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        start_row += 1

        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col_num, value=header)
            cell.alignment = wrap_alignment
            cell.border = thin_border
            cell.font = bold_font
        start_row += 1

        year_data = data.filter(year=year_value)
        for entry in year_data:
            cells = [
                worksheet.cell(row=start_row, column=1, value=entry.event_name),
                worksheet.cell(row=start_row, column=2, value=entry.num_participants),
                worksheet.cell(row=start_row, column=3,
                               value=f"{entry.date_from.strftime('%Y-%m-%d')} - {entry.date_to.strftime('%Y-%m-%d')}"),
                worksheet.cell(row=start_row, column=4, value=entry.activity_report_link),
            ]
            for cell in cells:
                cell.alignment = wrap_alignment
                cell.border = thin_border
                cell.font = bold_font
            start_row += 1

        start_row += 1

    virtual_workbook = BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response = HttpResponse(
        virtual_workbook,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=3_2_2_criteria_data.xlsx'
    return response


@login_required
def criteria_3_3_1(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='3_3_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="3_3_1")
    context = {
        'has_access': has_access,
        'users': users,
        '2021_records': Criteria_3_3_1.objects.filter(year = '2021'),
        '2022_records': Criteria_3_3_1.objects.filter(year = '2022'),
        '2023_records': Criteria_3_3_1.objects.filter(year = '2023'),
        '2024_records': Criteria_3_3_1.objects.filter(year = '2024'),
        '2025_records': Criteria_3_3_1.objects.filter(year = '2025')
    }
    return render(request, 'table/criteria_3_3_1.html', context=context)


@login_required
def criteria_3_3_1_form(request):
    if request.method == 'POST':
        form = CriteriaForm_3_3_1(request.POST)
        print(form.data)
        if form.is_valid():
            form.save() 
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_3_3_1')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_3_3_1()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_3_3_1.html', context=context)


@login_required
def upload_excel_3_3_1(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=3, values_only=True):
                    year = row[0]
                    title_of_paper = row[1]
                    authors = row[2]
                    department = row[3]
                    journal_name = row[4]
                    year_of_publication = row[5]
                    issn_number = row[6]
                    journal_website_link = row[7]
                    article_link = row[8]
                    is_ugc_care_listed = row[9]

                    Criteria_3_3_1.objects.create(
                        year=year,
                        title_of_paper=title_of_paper,
                        authors=authors,
                        department=department,
                        journal_name=journal_name,
                        year_of_publication=year_of_publication,
                        issn_number=issn_number,
                        journal_website_link=journal_website_link,
                        article_link=article_link,
                        is_ugc_care_listed = True if row[9] == 1 else False
                    )
            messages.success(request, 'Excel file uploaded and processed successfully.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
        return redirect('criteria_3_3_1')
    messages.error(request, 'Please upload a valid Excel file.')
    return redirect('criteria_3_3_1')


@login_required
def export_excel_3_3_1(request):
    data = Criteria_3_3_1.objects.all().order_by('year')

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Criteria 3.3.1"

    headers = [
        'Title of paper',
        'Name of the author/s',
        'Department of the teacher',
        'Name of journal',
        'Calendar Year of publication',
        'ISSN number',
        'Link to website of the Journal',
        'Link to article / paper / abstract of the article',
        'Is it listed in UGC Care list'
    ]

    column_width = 21.89
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = column_width

    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='top')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    bold_font = Font(bold=True)

    yellow_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

    title_cell = worksheet.cell(row=1, column=1,
                                value='3.3.1 Number of research papers published per teacher in the Journals notified '
                                      'on UGC CARE list during the last five years')

    start_row = 3

    for year in Criteria_3_3_1.YEAR_CHOICES:
        year_value, year_label = year

        worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
        year_cell = worksheet.cell(row=start_row, column=1, value=f"Year: {year_label}")
        year_cell.alignment = center_alignment
        year_cell.font = bold_font
        year_cell.fill = yellow_fill

        for row in worksheet.iter_rows(min_row=start_row, max_row=start_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        start_row += 1

        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col_num, value=header)
            cell.alignment = wrap_alignment
            cell.border = thin_border
            cell.font = bold_font
        start_row += 1

        year_data = data.filter(year=year_value)
        for entry in year_data:
            cells = [
                worksheet.cell(row=start_row, column=1, value=entry.title_of_paper),
                worksheet.cell(row=start_row, column=2, value=entry.authors),
                worksheet.cell(row=start_row, column=3, value=entry.department),
                worksheet.cell(row=start_row, column=4, value=entry.journal_name),
                worksheet.cell(row=start_row, column=5, value=entry.year_of_publication),
                worksheet.cell(row=start_row, column=6, value=entry.issn_number),
                worksheet.cell(row=start_row, column=7, value=entry.journal_website_link),
                worksheet.cell(row=start_row, column=8, value=entry.article_link),
                worksheet.cell(row=start_row, column=9, value=entry.is_ugc_care_listed),
            ]
            for cell in cells:
                cell.alignment = wrap_alignment
                cell.border = thin_border
                cell.font = bold_font
            start_row += 1

        start_row += 1

    virtual_workbook = BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response = HttpResponse(
        virtual_workbook,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=3_3_1_criteria_data.xlsx'
    return response


@login_required
def criteria_3_3_2(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='3_3_2').exists()
    users = UserCriteriaLink.objects.filter(criteria="3_3_2")
    context = {
        'has_access': has_access,
        'users': users,
        'records': Criteria_3_3_2.objects.all(),
    }
    return render(request, 'table/criteria_3_3_2.html', context=context)


@login_required
def criteria_3_3_2_form(request):
    if request.method == 'POST':
        form = CriteriaForm_3_3_2(request.POST)
        print(form.data)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_3_3_2')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_3_3_2()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_3_3_2.html', context=context)


@login_required
def upload_excel_3_3_2(request):
    pass


@login_required
def export_excel_3_3_2(request):
    pass


@login_required
def criteria_3_4_1_form(request):
    # Get the existing instance (if any)
    instance = Criteria_3_4_1.objects.first()

    if request.method == 'POST':
        form = CriteriaForm_3_4_1(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('criteria_3_4_1')
        else:
            print(form.errors)
    else:
        form = CriteriaForm_3_4_1(instance=instance)

    return render(request, 'form/criteria_3_4_1.html', {'form': form})


@login_required
def criteria_3_4_1(request):
    criteria_data = Criteria_3_4_1.objects.first()
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='3_4_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="3_4_1")
    context = {
        'criteria_data': criteria_data,
        'has_access': has_access,
        'users': users,
    }
    if not criteria_data:
        context['no_data'] = True
    return render(request, 'table/criteria_3_4_1.html', context)


def criteria_3_4_1_delete(request):
    Criteria_3_4_1.objects.all().delete()
    return redirect('criteria_3_4_1')


@login_required
def criteria_3_4_2_form(request):
    # Get the existing instance (if any)
    instance = Criteria_3_4_2.objects.first()

    if request.method == 'POST':
        form = CriteriaForm_3_4_2(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('criteria_3_4_2')
        else:
            print(form.errors)
    else:
        form = CriteriaForm_3_4_2(instance=instance)

    return render(request, 'form/criteria_3_4_2.html', {'form': form})


@login_required
def criteria_3_4_2(request):
    criteria_data = Criteria_3_4_2.objects.first()
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='3_4_2').exists()
    users = UserCriteriaLink.objects.filter(criteria="3_4_2")
    context = {
        'criteria_data': criteria_data,
        'has_access': has_access,
        'users': users,
    }
    if not criteria_data:
        context['no_data'] = True
    return render(request, 'table/criteria_3_4_2.html', context)


@login_required
def criteria_3_4_2_delete(request):
    Criteria_3_4_2.objects.all().delete()
    return redirect('criteria_3_4_2')


@login_required
def criteria_3_4_3(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='3_4_3').exists()
    users = UserCriteriaLink.objects.filter(criteria="3_4_3")
    context = {
        'has_access': has_access,
        'users': users,
        'records': Criteria_3_4_3.objects.all(),
    }
    return render(request, 'table/criteria_3_4_3.html', context=context)


@login_required
def criteria_3_4_3_form(request):
    if request.method == 'POST':
        form = CriteriaForm_3_4_3(request.POST)
        print(form.data)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_3_4_3')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_3_4_3()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_3_4_3.html', context=context)


@login_required
def upload_excel_3_4_3(request):
    pass


@login_required
def export_excel_3_4_3(request):
    pass


@login_required
def criteria_3_5_1(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='3_5_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="3_5_1")
    context = {
        'has_access': has_access,
        'users': users,
        '2021_records': Criteria_3_5_1.objects.filter(year = '2021'),
        '2022_records': Criteria_3_5_1.objects.filter(year = '2022'),
        '2023_records': Criteria_3_5_1.objects.filter(year = '2023'),
        '2024_records': Criteria_3_5_1.objects.filter(year = '2024'),
        '2025_records': Criteria_3_5_1.objects.filter(year = '2025')
    }
    return render(request, 'table/criteria_3_5_1.html', context=context)


@login_required
def criteria_3_5_1_form(request):
    if request.method == 'POST':
        form = CriteriaForm_3_5_1(request.POST)
        if form.is_valid():
            form.save() 
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_3_5_1')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_3_5_1()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_3_5_1.html', context=context)


@login_required
def upload_excel_3_5_1(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Assuming the Excel columns match the model fields
                    year = row[0]
                    mou_name = row[1]
                    institution_name = row[2]
                    signing_year = row[3]
                    purpose = row[4]
                    duration = row[5]
                    activities = row[6]
                    document_link = row[7]

                    # Creating the Criteria_3_5_1 object
                    Criteria_3_5_1.objects.create(
                        year=year,
                        mou_name=mou_name,
                        institution_name=institution_name,
                        signing_year=signing_year,
                        pourpose=purpose,
                        duration=duration,
                        activities=activities,
                        document_link=document_link
                    )

            messages.success(request, 'Excel file uploaded and processed successfully.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
        return redirect('criteria_3_5_1')
    messages.error(request, 'Please upload a valid Excel file.')
    return redirect('criteria_3_5_1')


@login_required
def export_excel_3_5_1(request):
    data = Criteria_3_5_1.objects.all().order_by('year')

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Criteria 3.5.1"

    headers = [
        'Sl. No.',
        'Name of the MoU / linkage',
        'Name of the institution / industry with whom the MoU / linkage is made, with contact details',
        'Year of signing MoU / linkage',
        'Purpose of the MoU/Linkage (internship, on-the-job training, project work, student / faculty exchange and '
        'collaborative research)',
        'Duration of MoU / linkage',
        'List the actual activities under each MoU/Linkage and web-links year-wise',
        'Link to the relevant document'
    ]

    column_width = 21.89
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = column_width

    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='top')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    bold_font = Font(bold=True)

    yellow_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

    title_cell = worksheet.cell(row=1, column=1,
                                value='3.5.1. Number of functional MoUs/linkages with institutions/ industries in '
                                      'India '
                                      'and abroad for internship, on-the-job training, project work, student / faculty'
                                      ' exchange and collaborative research  during the last five')

    start_row = 3

    for year in Criteria_3_5_1.YEAR_CHOICES:
        year_value, year_label = year

        worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
        year_cell = worksheet.cell(row=start_row, column=1, value=f"Year: {year_label}")
        year_cell.alignment = center_alignment
        year_cell.font = bold_font
        year_cell.fill = yellow_fill

        for row in worksheet.iter_rows(min_row=start_row, max_row=start_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        start_row += 1

        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col_num, value=header)
            cell.alignment = wrap_alignment
            cell.border = thin_border
            cell.font = bold_font
        start_row += 1

        year_data = data.filter(year=year_value)
        for entry in year_data:
            cells = [
                worksheet.cell(row=start_row, column=1, value=start_row - start_row + 1),  # Sl. No.
                worksheet.cell(row=start_row, column=2, value=entry.mou_name),
                worksheet.cell(row=start_row, column=3, value=entry.institution_name),
                worksheet.cell(row=start_row, column=4, value=entry.signing_year),
                worksheet.cell(row=start_row, column=5, value=entry.pourpose),
                worksheet.cell(row=start_row, column=6, value=entry.duration),
                worksheet.cell(row=start_row, column=7, value=entry.activities),
                worksheet.cell(row=start_row, column=8, value=entry.document_link),
            ]
            for cell in cells:
                cell.alignment = wrap_alignment
                cell.border = thin_border
                cell.font = bold_font
            start_row += 1

        start_row += 1

    virtual_workbook = BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response = HttpResponse(
        virtual_workbook,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=3_5_1_criteria_data.xlsx'
    return response

