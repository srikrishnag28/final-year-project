from django.shortcuts import render, redirect, HttpResponse
from .models import *
from .forms import *
from datetime import datetime
from django.db import transaction
import openpyxl
from openpyxl import Workbook
from io import BytesIO
from django.contrib import messages
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from user.models import UserCriteriaLink
from django.contrib.auth.decorators import login_required
# Create your views here.


@login_required
def criteria_1_1_1_form(request):
    # Get the existing instance (if any)
    instance = Criteria_1_1_1.objects.first()

    if request.method == 'POST':
        form = CriteriaForm_1_1_1(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('criteria_1_1_1')
        else:
            print(form.errors)
    else:
        form = CriteriaForm_1_1_1(instance=instance)

    return render(request, 'form/criteria_1_1_1.html', {'form': form})


@login_required
def criteria_1_1_1(request):
    criteria_data = Criteria_1_1_1.objects.first()
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='1_1_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="1_1_1")
    context = {
        'criteria_data': criteria_data,
        'has_access': has_access,
        'users': users,
    }
    if not criteria_data:
        context['no_data'] = True
    return render(request, 'table/criteria_1_1_1.html', context)


@login_required
def criteria_1_1_1_delete(request):
    Criteria_1_1_1.objects.all().delete()
    return redirect('criteria_1_1_1')


@login_required
def criteria_1_2_2(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='1_2_2').exists()
    users = UserCriteriaLink.objects.filter(criteria="1_2_2")
    context = {
        'has_access': has_access,
        'users': users,
        '2021_records': Criteria_1_2_2.objects.filter(year = '2021'),
        '2022_records': Criteria_1_2_2.objects.filter(year = '2022'),
        '2023_records': Criteria_1_2_2.objects.filter(year = '2023'),
        '2024_records': Criteria_1_2_2.objects.filter(year = '2024'),
        '2025_records': Criteria_1_2_2.objects.filter(year = '2025')
    }
    return render(request, 'table/criteria_1_2_2.html', context=context)


@login_required
def criteria_1_2_2_form(request):
    if request.method == 'POST':
        form = CriteriaForm_1_2_2(request.POST)
        print(form)
        if form.is_valid():
            form.save() 
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_1_2_2')
        else:
            print(form.errors)
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_1_2_2()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_1_2_2.html', context=context)


@login_required
def upload_excel_1_2_2(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    year = row[0]
                    course_name = row[1]
                    course_code = row[2]
                    year_of_study = row[3]
                    period_from = row[4]  
                    period_to = row[5]  
                    duration = row[6]
                    students_enrolled = row[7]
                    students_completed = row[8]

                    Criteria_1_2_2.objects.create(
                        year=year,
                        course_name=course_name,
                        course_code=course_code,
                        year_of_study=year_of_study,
                        period_from=period_from,
                        period_to=period_to,
                        duration=duration,
                        students_enrolled=students_enrolled,
                        students_completed=students_completed,
                    )
            messages.success(request, 'Excel File added successfully')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
        return redirect('criteria_1_2_2')
    messages.error(request, 'Upload a valid Excel file')
    return redirect('criteria_1_2_2')


@login_required
def export_excel_1_2_2(request):
    data = Criteria_1_2_2.objects.all().order_by('year')

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Criteria 1.2.2"

    headers = [
        'Name of Certificate/Value added course offered and online courses of MOOCs, SWAYAM, NPTEL etc. where the students of the institution have enrolled and successfully completed',
        'Course Code (if any)',
        'Year of offering/study',
        'Period (from date - to date)',
        'Duration of course',
        'Number of students enrolled in the year',
        'Number of Students completing the course in the year'
    ]

    column_width = 21.89
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = column_width

    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='top')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    bold_font = Font(bold=True)

    yellow_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

    title_cell = worksheet.cell(row=1, column=1, value='1.2.1 Number of Certificate/Value added courses offered and '
                                                       'online courses of MOOCs, SWAYAM, NPTEL etc.  where the students'
                                                       ' of the institution have enrolled and successfully completed'
                                                       ' during the last five years)')

    title_cell_2 = worksheet.cell(row=2, column=1, value='1.2.2 Percentage of students enrolled in Certificate/ Value'
                                                         ' added courses and also completed online courses of MOOCs, '
                                                         'SWAYAM, NPTEL etc.')

    start_row = 4

    for year in Criteria_1_2_2.YEAR_CHOICES:
        year_value, year_label = year

        worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
        year_cell = worksheet.cell(row=start_row, column=1, value=f"Year: {year_label}")
        year_cell.alignment = center_alignment
        year_cell.font = bold_font
        year_cell.fill = yellow_fill

        for row in worksheet.iter_rows(min_row=start_row, max_row=start_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        start_row += 1

        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col_num, value=header)
            cell.alignment = wrap_alignment
            cell.border = thin_border
            cell.font = bold_font
        start_row += 1

        year_data = data.filter(year=year_value)
        for entry in year_data:
            cells = [
                worksheet.cell(row=start_row, column=1, value=entry.course_name),
                worksheet.cell(row=start_row, column=2, value=entry.course_code),
                worksheet.cell(row=start_row, column=3, value=entry.year_of_study),
                worksheet.cell(row=start_row, column=4, value=f"{entry.period_from.strftime('%Y-%m-%d')} - {entry.period_to.strftime('%Y-%m-%d')}"),
                worksheet.cell(row=start_row, column=5, value=entry.duration),
                worksheet.cell(row=start_row, column=6, value=entry.students_enrolled),
                worksheet.cell(row=start_row, column=7, value=entry.students_completed),
            ]
            for cell in cells:
                cell.alignment = wrap_alignment
                cell.border = thin_border
                cell.font = bold_font
            start_row += 1

        start_row += 1

    virtual_workbook = BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response = HttpResponse(
        virtual_workbook,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=1_2_2_criteria_data.xlsx'
    return response


@login_required
def criteria_1_3_1_form(request):
    # Get the existing instance (if any)
    instance = Criteria_1_3_1.objects.first()

    if request.method == 'POST':
        form = CriteriaForm_1_3_1(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('criteria_1_3_1')
        else:
            print(form.errors)
    else:
        form = CriteriaForm_1_3_1(instance=instance)

    return render(request, 'form/criteria_1_3_1.html', {'form': form})


@login_required
def criteria_1_3_1(request):
    criteria_data = Criteria_1_3_1.objects.first()
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='1_3_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="1_3_1")
    context = {
        'criteria_data': criteria_data,
        'has_access': has_access,
        'users': users,
    }
    if not criteria_data:
        context['no_data'] = True
    return render(request, 'table/criteria_1_3_1.html', context)


@login_required
def criteria_1_3_1_delete(request):
    Criteria_1_3_1.objects.all().delete()
    return redirect('criteria_1_3_1')


@login_required
def criteria_1_3_2(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='1_3_2').exists()
    users = UserCriteriaLink.objects.filter(criteria="1_3_2")
    context = {
        'users': users,
        'has_access': has_access,
        'records': Criteria_1_3_2.objects.all(),
    }
    return render(request, 'table/criteria_1_3_2.html', context=context)


@login_required
def criteria_1_3_2_form(request):
    if request.method == 'POST':
        form = CriteriaForm_1_3_2(request.POST)
        print(form.data)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_1_3_2')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_1_3_2()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_1_3_2.html', context=context)


@login_required
def upload_excel_1_3_2(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    programme_name = row[0]
                    programme_code = row[1]
                    student_list = row[2]
                    document_link = row[3]

                    Criteria_1_3_2.objects.create(
                        programme_name=programme_name,
                        programme_code=programme_code,
                        student_list=student_list,
                        document_link=document_link
                    )
            messages.success(request, 'Excel file uploaded and processed successfully.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
        return redirect('criteria_1_3_2')
    messages.error(request, 'Please upload a valid Excel file.')
    return redirect('criteria_1_3_2')


@login_required
def export_excel_1_3_2(request):
    data = Criteria_1_3_2.objects.all()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Criteria 1.3.2"

    headers = [
        'Programme name',
        'Program Code',
        'List of students undertaking project work/field work/internship',
        'Link to the relevant document'
    ]

    column_width = 25  # Adjust column width for readability
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = column_width

    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    bold_font = Font(bold=True)
    yellow_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

    # Adding title cell
    title_cell = worksheet.cell(row=1, column=1, value='1.3.2 Percentage  of students undertaking project work/field '
                                                       'work/internship (Data for the latest completed academic year)')
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    # Adding headers
    header_row = 3
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=header_row, column=col_num, value=header)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        cell.font = bold_font
        cell.fill = yellow_fill

    # Adding data
    start_row = 4
    for entry in data:
        cells = [
            worksheet.cell(row=start_row, column=1, value=entry.programme_name),
            worksheet.cell(row=start_row, column=2, value=entry.programme_code),
            worksheet.cell(row=start_row, column=3, value=entry.student_list),
            worksheet.cell(row=start_row, column=4, value=entry.document_link),
        ]
        for cell in cells:
            cell.alignment = wrap_alignment
            cell.border = thin_border
        start_row += 1

    # Prepare the response
    virtual_workbook = BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response = HttpResponse(
        virtual_workbook,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=criteria_1_3_2.xlsx'
    return response
