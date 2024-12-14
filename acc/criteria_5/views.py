from django.shortcuts import render, redirect, HttpResponse
from .models import *
from .forms import *
from datetime import datetime
from django.db import transaction
import openpyxl
from openpyxl import Workbook
from io import BytesIO
from django.contrib import messages
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from user.models import UserCriteriaLink
from django.contrib.auth.decorators import login_required
# Create your views here.


@login_required
def criteria_5_1_1(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='5_1_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="5_1_1")
    context = {
        'has_access': has_access,
        'users': users,
        'records': Criteria_5_1_1.objects.all(),
    }
    return render(request, 'table/criteria_5_1_1.html', context=context)


@login_required
def criteria_5_1_1_form(request):
    if request.method == 'POST':
        form = CriteriaForm_5_1_1(request.POST)
        print(form.data)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_5_1_1')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_5_1_1()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_5_1_1.html', context=context)


@login_required
def upload_excel_5_1_1(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Extracting data from the Excel row
                    year = row[0]
                    scheme_name = row[1]
                    government_or_non_government = row[2]
                    individual_or_organization_name = row[3]
                    number_of_students_benefited = row[4]
                    amount_in_inr = row[5]

                    # Validation for GOVERNMENT_CHOICES
                    if government_or_non_government not in dict(Criteria_5_1_1.GOVERNMENT_CHOICES):
                        raise ValueError(
                            f"Invalid value for government_or_non_government: {government_or_non_government}. "
                            f"Accepted values are {list(dict(Criteria_5_1_1.GOVERNMENT_CHOICES).keys())}."
                        )

                    # Creating the Criteria_5_1_1 object
                    Criteria_5_1_1.objects.create(
                        year=year,
                        scheme_name=scheme_name,
                        government_or_non_government=government_or_non_government,
                        individual_or_organization_name=individual_or_organization_name,
                        number_of_students_benefited=number_of_students_benefited,
                        amount_in_inr=amount_in_inr
                    )
            messages.success(request, 'Excel file uploaded and processed successfully.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
        return redirect('criteria_5_1_1')
    messages.error(request, 'Please upload a valid Excel file.')
    return redirect('criteria_5_1_1')


@login_required
def export_excel_5_1_1(request):
    data = Criteria_5_1_1.objects.all()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Criteria 5.1.1"

    headers = [
        'Year',
        'Name of the scheme',
        'Government/Non-government',
        'Name of the individual/organisation',
        'Number of students benefited',
        'Amount (INR)'
    ]

    column_width = 25  # Adjust column width for readability
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = column_width

    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    bold_font = Font(bold=True)
    yellow_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

    # Adding title cell
    title_cell = worksheet.cell(row=1, column=1, value='5.1.1 Percentage of students benefited by scholarships and freeships provided by the institution, government and non-government bodies, industries, individuals, philanthropists during the last five years')
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    # Adding headers
    header_row = 3
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=header_row, column=col_num, value=header)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        cell.font = bold_font
        cell.fill = yellow_fill

    # Adding data
    start_row = 4
    for entry in data:
        cells = [
            worksheet.cell(row=start_row, column=1, value=entry.year),
            worksheet.cell(row=start_row, column=2, value=entry.scheme_name),
            worksheet.cell(row=start_row, column=3, value=entry.government_or_non_government),
            worksheet.cell(row=start_row, column=4, value=entry.individual_or_organization_name),
            worksheet.cell(row=start_row, column=5, value=entry.number_of_students_benefited),
            worksheet.cell(row=start_row, column=6, value=entry.amount_in_inr),
        ]
        for cell in cells:
            cell.alignment = wrap_alignment
            cell.border = thin_border
        start_row += 1

    # Prepare the response
    virtual_workbook = BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response = HttpResponse(
        virtual_workbook,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=criteria_5_1_1.xlsx'
    return response


@login_required
def criteria_5_1_2(request):
    criteria_data = Criteria_5_1_2.objects.first()
    flag = UserCriteriaLink.objects.filter(user=request.user, criteria='5_1_2').exists()
    if request.method == 'POST':
        form = CriteriaForm_5_1_2(request.POST)
        if form.is_valid():
            if criteria_data:
                form.instance.pk = criteria_data.pk
            form.save()
            messages.success(request, 'Data Added/Updated Successfully')
            return redirect('criteria_5_1_2')
        else:
            messages.error(request, 'Please enter valid data.')
    form = CriteriaForm_5_1_2(instance=criteria_data) if criteria_data else CriteriaForm_5_1_2()
    context = {
        'form': form,
        'criteria_data': criteria_data,
        'flag': flag,
    }
    return render(request, 'table/criteria_5_1_2.html', context)


@login_required
def criteria_5_1_3(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='5_1_3').exists()
    users = UserCriteriaLink.objects.filter(criteria="5_1_3")
    context = {
        'has_access': has_access,
        'users': users,
        'records': Criteria_5_1_3.objects.all(),
    }
    return render(request, 'table/criteria_5_1_3.html', context=context)


@login_required
def criteria_5_1_3_form(request):
    if request.method == 'POST':
        form = CriteriaForm_5_1_3(request.POST)
        print(form.data)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_5_1_3')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_5_1_3()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_5_1_3.html', context=context)


@login_required
def upload_excel_5_1_3(request):
    pass


@login_required
def export_excel_5_1_3(request):
    pass


@login_required
def criteria_5_1_4(request):
    criteria_data = Criteria_5_1_4.objects.first()
    flag = UserCriteriaLink.objects.filter(user=request.user, criteria='5_1_4').exists()
    if request.method == 'POST':
        form = CriteriaForm_5_1_4(request.POST)
        if form.is_valid():
            if criteria_data:
                form.instance.pk = criteria_data.pk
            form.save()
            messages.success(request, 'Data Added/Updated Successfully')
            return redirect('criteria_5_1_4')
        else:
            messages.error(request, 'Please enter valid data.')
    form = CriteriaForm_5_1_4(instance=criteria_data) if criteria_data else CriteriaForm_5_1_4()
    context = {
        'form': form,
        'criteria_data': criteria_data,
        'flag': flag,
    }
    return render(request, 'table/criteria_5_1_4.html', context)


@login_required
def criteria_5_2_1(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='5_2_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="5_2_1")
    context = {
        'has_access': has_access,
        'users': users,
        '1_records': Criteria_5_2_1_1.objects.all(),
        '2_records': Criteria_5_2_1_2.objects.all(),
    }
    return render(request, 'table/criteria_5_2_1.html', context=context)


@login_required
def criteria_5_2_1_1_form(request):
    if request.method == 'POST':
        form = CriteriaForm_5_2_1_1(request.POST)
        print(form.data)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_5_2_1')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_5_2_1_1()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_5_2_1_1.html', context=context)


@login_required
def criteria_5_2_1_2_form(request):
    if request.method == 'POST':
        form = CriteriaForm_5_2_1_2(request.POST)
        print(form.data)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_5_2_1')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_5_2_1_2()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_5_2_1_2.html', context=context)


@login_required
def upload_excel_5_2_1_1(request):
    pass


@login_required
def upload_excel_5_2_1_2(request):
    pass


@login_required
def export_excel_5_2_1(request):
    pass


@login_required
def criteria_5_2_2(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='5_2_2').exists()
    users = UserCriteriaLink.objects.filter(criteria="5_2_2")
    context = {
        'has_access': has_access,
        'users': users,
        'records': Criteria_5_2_2.objects.all(),
    }
    return render(request, 'table/criteria_5_2_2.html', context=context)


@login_required
def criteria_5_2_2_form(request):
    if request.method == 'POST':
        form = CriteriaForm_5_2_2(request.POST)
        print(form.data)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_5_2_2')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_5_2_2()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_5_2_2.html', context=context)


@login_required
def upload_excel_5_2_2(request):
    pass


@login_required
def export_excel_5_2_2(request):
    pass


@login_required
def criteria_5_3_1(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='5_3_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="5_3_1")
    context = {
        'has_access': has_access,
        'users': users,
        'records': Criteria_5_3_1.objects.all(),
    }
    return render(request, 'table/criteria_5_3_1.html', context=context)


@login_required
def criteria_5_3_1_form(request):
    if request.method == 'POST':
        form = CriteriaForm_5_3_1(request.POST)
        print(form.data)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_5_3_1')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_5_3_1()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_5_3_1.html', context=context)


@login_required
def upload_excel_5_3_1(request):
    pass


@login_required
def export_excel_5_3_1(request):
    pass


@login_required
def criteria_5_3_2(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='5_3_2').exists()
    users = UserCriteriaLink.objects.filter(criteria="5_3_2")
    context = {
        'has_access': has_access,
        'users': users,
        '2021_records': Criteria_5_3_2.objects.filter(year = '2021'),
        '2022_records': Criteria_5_3_2.objects.filter(year = '2022'),
        '2023_records': Criteria_5_3_2.objects.filter(year = '2023'),
        '2024_records': Criteria_5_3_2.objects.filter(year = '2024'),
        '2025_records': Criteria_5_3_2.objects.filter(year = '2025')
    }
    return render(request, 'table/criteria_5_3_2.html', context=context)


@login_required
def criteria_5_3_2_form(request):
    if request.method == 'POST':
        form = CriteriaForm_5_3_2(request.POST)
        if form.is_valid():
            form.save() 
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_5_3_2')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_5_3_2()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_5_3_2.html', context=context)


@login_required
def upload_excel_5_3_2(request):
    pass


@login_required
def export_excel_5_3_2(request):
    pass


@login_required
def criteria_5_4_1_form(request):
    # Get the existing instance (if any)
    instance = Criteria_5_4_1.objects.first()

    if request.method == 'POST':
        form = CriteriaForm_5_4_1(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('criteria_5_4_1')
        else:
            print(form.errors)
    else:
        form = CriteriaForm_5_4_1(instance=instance)

    return render(request, 'form/criteria_5_4_1.html', {'form': form})


@login_required
def criteria_5_4_1(request):
    criteria_data = Criteria_5_4_1.objects.first()
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='5_4_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="5_4_1")
    context = {
        'criteria_data': criteria_data,
        'has_access': has_access,
        'users': users,
    }
    if not criteria_data:
        context['no_data'] = True
    return render(request, 'table/criteria_5_4_1.html', context)


@login_required
def criteria_5_4_1_delete(request):
    Criteria_5_4_1.objects.all().delete()
    return redirect('criteria_5_4_1')