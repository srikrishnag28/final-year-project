from django.shortcuts import render, HttpResponse, redirect
from .models import Criteria_2_1_1, Criteria_2_7_1
from .forms import *
from datetime import datetime
from django.db import transaction
import openpyxl
from openpyxl import Workbook
from io import BytesIO
from django.contrib import messages
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from user.models import UserCriteriaLink
from django.contrib.auth.decorators import login_required


@login_required
def criteria_2_3_1_form(request):
    # Get the existing instance (if any)
    instance = Criteria_2_3_1.objects.first()

    if request.method == 'POST':
        form = CriteriaForm_2_3_1(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('criteria_2_3_1')
        else:
            print(form.errors)
    else:
        form = CriteriaForm_2_3_1(instance=instance)

    return render(request, 'form/criteria_2_3_1.html', {'form': form})


@login_required
def criteria_2_3_1(request):
    criteria_data = Criteria_2_3_1.objects.first()
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='2_3_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="2_3_1")
    context = {
        'criteria_data': criteria_data,
        'has_access': has_access,
        'users': users,
    }
    if not criteria_data:
        context['no_data'] = True
    return render(request, 'table/criteria_2_3_1.html', context)


@login_required
def criteria_2_3_1_delete(request):
    Criteria_2_3_1.objects.all().delete()
    return redirect('criteria_2_3_1')

@login_required
def criteria_2_4_2(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='2_4_2').exists()
    users = UserCriteriaLink.objects.filter(criteria="2_4_2")
    context = {
        'has_access': has_access,
        'users': users,
        '1_records': Criteria_2_4_2_1.objects.all(),
        '2_records': Criteria_2_4_2_2.objects.all(),
    }
    return render(request, 'table/criteria_2_4_2.html', context=context)


@login_required
def criteria_2_4_2_1_form(request):
    if request.method == 'POST':
        form = CriteriaForm_2_4_2_1(request.POST)
        print(form.data)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_2_4_2')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_2_4_2_1()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_2_4_2_1.html', context=context)


@login_required
def criteria_2_4_2_2_form(request):
    if request.method == 'POST':
        form = CriteriaForm_2_4_2_2(request.POST)
        print(form.data)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_2_4_2')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_2_4_2_2()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_2_4_2_2.html', context=context)


@login_required
def upload_excel_2_4_2_1(request):
    pass


@login_required
def upload_excel_2_4_2_2(request):
    pass


def export_excel_2_4_2(request):
    pass


@login_required
def criteria_2_5_1_form(request):
    # Get the existing instance (if any)
    instance = Criteria_2_5_1.objects.first()

    if request.method == 'POST':
        form = CriteriaForm_2_5_1(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('criteria_2_5_1')
        else:
            print(form.errors)
    else:
        form = CriteriaForm_2_5_1(instance=instance)

    return render(request, 'form/criteria_2_5_1.html', {'form': form})


@login_required
def criteria_2_5_1(request):
    criteria_data = Criteria_2_5_1.objects.first()
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='2_5_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="2_5_1")
    context = {
        'criteria_data': criteria_data,
        'has_access': has_access,
        'users': users,
    }
    if not criteria_data:
        context['no_data'] = True
    return render(request, 'table/criteria_2_5_1.html', context)


@login_required
def criteria_2_5_1_delete(request):
    Criteria_2_5_1.objects.all().delete()
    return redirect('criteria_2_5_1')


@login_required
def criteria_2_6_1_form(request):
    # Get the existing instance (if any)
    instance = Criteria_2_6_1.objects.first()

    if request.method == 'POST':
        form = CriteriaForm_2_6_1(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('criteria_2_6_1')
        else:
            print(form.errors)
    else:
        form = CriteriaForm_2_6_1(instance=instance)

    return render(request, 'form/criteria_2_6_1.html', {'form': form})


@login_required
def criteria_2_6_1(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='2_6_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="2_6_1")
    context = {
        'criteria_data': criteria_data,
        'has_access': has_access,
        'users': users,
    }
    if not criteria_data:
        context['no_data'] = True
    return render(request, 'table/criteria_2_6_1.html', context)


@login_required
def criteria_2_6_1_delete(request):
    Criteria_2_6_1.objects.all().delete()
    return redirect('criteria_2_6_1')


@login_required
def criteria_2_6_2_form(request):
    # Get the existing instance (if any)
    instance = Criteria_2_6_2.objects.first()

    if request.method == 'POST':
        form = CriteriaForm_2_6_2(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('criteria_2_6_2')
        else:
            print(form.errors)
    else:
        form = CriteriaForm_2_6_2(instance=instance)

    return render(request, 'form/criteria_2_6_2.html', {'form': form})


@login_required
def criteria_2_6_2(request):
    criteria_data = Criteria_2_6_2.objects.first()
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='2_6_2').exists()
    users = UserCriteriaLink.objects.filter(criteria="2_6_2")
    context = {
        'criteria_data': criteria_data,
        'has_access': has_access,
        'users': users,
    }
    if not criteria_data:
        context['no_data'] = True
    return render(request, 'table/criteria_2_6_2.html', context)


@login_required
def criteria_2_6_2_delete(request):
    Criteria_2_6_2.objects.all().delete()
    return redirect('criteria_2_6_2')


@login_required
def criteria_2_6_3(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='2_6_3').exists()
    users = UserCriteriaLink.objects.filter(criteria="2_6_3")
    context = {
        'has_access': has_access,
        'users': users,
        'records': Criteria_2_6_3.objects.all(),
    }
    return render(request, 'table/criteria_2_6_3.html', context=context)


@login_required
def criteria_2_6_3_form(request):
    if request.method == 'POST':
        form = CriteriaForm_2_6_3(request.POST)
        print(form.data)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_2_6_3')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_2_6_3()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_2_6_3.html', context=context)


@login_required
def upload_excel_2_6_3(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    year = row[0]
                    program_code = row[1]
                    program_name = row[2]
                    number_of_students_appeared = row[3]
                    number_of_students_passed = row[4]

                    Criteria_2_6_3.objects.create(
                        year=year,
                        program_code=program_code,
                        program_name=program_name,
                        number_of_students_appeared=number_of_students_appeared,
                        number_of_students_passed=number_of_students_passed
                    )
            messages.success(request, 'Excel file uploaded and processed successfully.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
        return redirect('criteria_2_6_3')
    messages.error(request, 'Please upload a valid Excel file.')
    return redirect('criteria_2_6_3')


@login_required
def export_excel_2_6_3(request):
    data = Criteria_2_6_3.objects.all()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Criteria 2.6.3"

    headers = [
        'Year',
        'Program Code',
        'Program Name',
        'Number of students appeared in the final year examination',
        'Number of students passed in final year examination'
    ]

    column_width = 25  # Adjust column width for readability
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = column_width

    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    bold_font = Font(bold=True)
    yellow_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

    # Adding title cell
    title_cell = worksheet.cell(row=1, column=1, value='2.6.3 Pass percentage of Students during last five years '
                                                       ' (excluding backlog students)')
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    # Adding headers
    header_row = 3
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=header_row, column=col_num, value=header)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        cell.font = bold_font
        cell.fill = yellow_fill

    # Adding data
    start_row = 4
    for entry in data:
        cells = [
            worksheet.cell(row=start_row, column=1, value=entry.year),
            worksheet.cell(row=start_row, column=2, value=entry.program_code),
            worksheet.cell(row=start_row, column=3, value=entry.program_name),
            worksheet.cell(row=start_row, column=4, value=entry.number_of_students_appeared),
            worksheet.cell(row=start_row, column=5, value=entry.number_of_students_passed),
        ]
        for cell in cells:
            cell.alignment = wrap_alignment
            cell.border = thin_border
        start_row += 1

    # Prepare the response
    virtual_workbook = BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response = HttpResponse(
        virtual_workbook,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=criteria_2_6_3.xlsx'
    return response


@login_required
def criteria_2_7_1(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='2_7_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="2_7_1")
    context = {
        'has_access': has_access,
        'users': users,
        'records': Criteria_2_7_1.objects.all(),
    }
    return render(request, 'table/criteria_2_7_1.html', context=context)


@login_required
def criteria_2_7_1_form(request):
    if request.method == 'POST':
        form = CriteriaForm_2_7_1(request.POST)
        print(form.data)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_2_7_1')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_2_7_1()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_2_7_1.html', context=context)


@login_required
def upload_excel_2_7_1(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    name = row[0]
                    gender = row[1]
                    category = row[2]
                    state_of_domicile = row[3]
                    nationality = row[4]
                    email = row[5]
                    program_name = row[6]
                    enrolment_id = row[7]
                    year_of_joining = row[8]

                    Criteria_2_7_1.objects.create(
                        name=name,
                        gender=gender,
                        category=category,
                        state_of_domicile=state_of_domicile,
                        nationality=nationality,
                        email=email,
                        program_name=program_name,
                        enrolment_id=enrolment_id,
                        year_of_joining=year_of_joining
                    )
            messages.success(request, 'Excel file uploaded and processed successfully.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
        return redirect('criteria_2_7_1')
    messages.error(request, 'Please upload a valid Excel file.')
    return redirect('criteria_2_7_1')


@login_required
def export_excel_2_7_1(request):
    data = Criteria_2_7_1.objects.all()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Criteria 2.7.1"

    headers = [
        'Name of the student',
        'Gender',
        'Category',
        'State of Domicile',
        'Nationality if other than Indian',
        'Email ID',
        'Program name',
        'Unique Enrolment ID / College ID / University enrolment number',
        'Year of joining'
    ]

    column_width = 25  # Adjust column width for readability
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = column_width

    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    bold_font = Font(bold=True)
    yellow_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

    # Adding title cell
    title_cell = worksheet.cell(row=1, column=1, value='2.7.1 Online student satisfaction survey regard to teaching '
                                                       'learning process (all currently enrolled students). details of '
                                                       'the students in the format mentioned below should be uploaded')
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    # Adding headers
    header_row = 3
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=header_row, column=col_num, value=header)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        cell.font = bold_font
        cell.fill = yellow_fill

    # Adding data
    start_row = 4
    for entry in data:
        cells = [
            worksheet.cell(row=start_row, column=1, value=entry.name),
            worksheet.cell(row=start_row, column=2, value=entry.gender),
            worksheet.cell(row=start_row, column=3, value=entry.category),
            worksheet.cell(row=start_row, column=4, value=entry.state_of_domicile),
            worksheet.cell(row=start_row, column=5, value=entry.nationality),
            worksheet.cell(row=start_row, column=6, value=entry.email),
            worksheet.cell(row=start_row, column=7, value=entry.program_name),
            worksheet.cell(row=start_row, column=8, value=entry.enrolment_id),
            worksheet.cell(row=start_row, column=9, value=entry.year_of_joining),
        ]
        for cell in cells:
            cell.alignment = wrap_alignment
            cell.border = thin_border
        start_row += 1

    # Prepare the response
    virtual_workbook = BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response = HttpResponse(
        virtual_workbook,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=criteria_2_7_1.xlsx'
    return response

