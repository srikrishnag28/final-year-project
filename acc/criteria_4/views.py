from django.shortcuts import render, redirect, HttpResponse
from .models import *
from .forms import *
from datetime import datetime
from django.db import transaction
import openpyxl
from openpyxl import Workbook
from io import BytesIO
from django.contrib import messages
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from user.models import UserCriteriaLink
from django.db.models import Sum
from django.contrib.auth.decorators import login_required
# Create your views here.


@login_required
def criteria_4_1_1_form(request):
    # Get the existing instance (if any)
    instance = Criteria_4_1_1.objects.first()

    if request.method == 'POST':
        form = CriteriaForm_4_1_1(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('criteria_4_1_1')
        else:
            print(form.errors)
    else:
        form = CriteriaForm_4_1_1(instance=instance)

    return render(request, 'form/criteria_4_1_1.html', {'form': form})


@login_required
def criteria_4_1_1(request):
    criteria_data = Criteria_4_1_1.objects.first()
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='4_1_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="4_1_1")
    context = {
        'criteria_data': criteria_data,
        'has_access': has_access,
        'users': users,
    }
    if not criteria_data:
        context['no_data'] = True
    return render(request, 'table/criteria_4_1_1.html', context)


@login_required
def criteria_4_1_1_delete(request):
    Criteria_4_1_1.objects.all().delete()
    return redirect('criteria_4_1_1')


@login_required
def criteria_4_1_2(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='4_1_2').exists()
    users = UserCriteriaLink.objects.filter(criteria="4_1_2")

    records_by_year = {}
    for year in [2021, 2022, 2023, 2024, 2025]:
        records = Criteria_4_1_2.objects.filter(year=year)
        total_amount = records.aggregate(total_amount=Sum('amount_in_lakhs'))['total_amount'] or 0
        records_by_year[f'{year}_records'] = {
            'records': records,
            'total': total_amount
        }
    context = {
        'has_access': has_access,
        'users': users,
        **records_by_year,
    }
    return render(request, 'table/criteria_4_1_2.html', context=context)


@login_required
def criteria_4_1_2_form(request):
    if request.method == 'POST':
        form = CriteriaForm_4_1_2(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data Added Successfully')
            return redirect('criteria_4_1_2')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_4_1_2()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_4_1_2.html', context=context)


@login_required
def upload_excel_4_1_2(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Assuming the Excel columns match the model fields
                    year = row[0]
                    head_of_expenditure = row[1]
                    item_of_expenditure = row[2]
                    amount_in_lakhs = row[3]

                    # Creating the Criteria_4_1_2 object
                    Criteria_4_1_2.objects.create(
                        year=year,
                        head_of_expenditure=head_of_expenditure,
                        item_of_expenditure=item_of_expenditure,
                        amount_in_lakhs=amount_in_lakhs
                    )

            messages.success(request, 'Excel file uploaded and processed successfully.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
        return redirect('criteria_4_1_2')
    messages.error(request, 'Please upload a valid Excel file.')
    return redirect('criteria_4_1_2')


@login_required
def export_excel_4_1_2(request):
    data = Criteria_4_1_2.objects.all().order_by('year')

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Criteria 4.1.2"

    headers = [
        'Head of expenditure (for example, capital expenditure)',
        'Item of expenditure (for example, construction of building, purchase of new equipment, '
        'furniture and fixtures, etc.)',
        'Amount (INR in Lakhs)'
    ]

    column_width = 21.89
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = column_width

    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='top')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    bold_font = Font(bold=True)

    yellow_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

    title_cell = worksheet.cell(row=1, column=1, value='4.1.2 Percentage of expenditure for infrastructure development'
                                                       ' and  augmentation excluding salary during the last five years')

    start_row = 3

    for year in Criteria_4_1_2.YEAR_CHOICES:
        year_value, year_label = year

        worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
        year_cell = worksheet.cell(row=start_row, column=1, value=f"Year: {year_label}")
        year_cell.alignment = center_alignment
        year_cell.font = bold_font
        year_cell.fill = yellow_fill

        for row in worksheet.iter_rows(min_row=start_row, max_row=start_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        start_row += 1

        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col_num, value=header)
            cell.alignment = wrap_alignment
            cell.border = thin_border
            cell.font = bold_font
        start_row += 1

        year_data = data.filter(year=year_value)
        for entry in year_data:
            cells = [
                worksheet.cell(row=start_row, column=1, value=entry.head_of_expenditure),
                worksheet.cell(row=start_row, column=2, value=entry.item_of_expenditure),
                worksheet.cell(row=start_row, column=3, value=entry.amount_in_lakhs),
            ]
            for cell in cells:
                cell.alignment = wrap_alignment
                cell.border = thin_border
                cell.font = bold_font
            start_row += 1
        worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
        summary_cell = worksheet.cell(row=start_row, column=1, value="Total")
        summary_amount_cell = worksheet.cell(row=start_row, column=3,
                                             value=f"=SUM(C{start_row - len(year_data)}:C{start_row - 1})")

        summary_cell.alignment = wrap_alignment
        summary_cell.border = thin_border
        summary_cell.font = bold_font
        summary_amount_cell.alignment = wrap_alignment
        summary_amount_cell.border = thin_border
        summary_amount_cell.font = bold_font
        start_row += 2

    virtual_workbook = BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response = HttpResponse(
        virtual_workbook,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=4_1_2_criteria_data.xlsx'
    return response


@login_required
def criteria_4_2_1_form(request):
    # Get the existing instance (if any)
    instance = Criteria_4_2_1.objects.first()

    if request.method == 'POST':
        form = CriteriaForm_4_2_1(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('criteria_4_2_1')
        else:
            print(form.errors)
    else:
        form = CriteriaForm_4_2_1(instance=instance)

    return render(request, 'form/criteria_4_2_1.html', {'form': form})


@login_required
def criteria_4_2_1(request):
    criteria_data = Criteria_4_2_1.objects.first()
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='4_2_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="4_2_1")
    context = {
        'criteria_data': criteria_data,
        'has_access': has_access,
        'users': users,
    }
    if not criteria_data:
        context['no_data'] = True
    return render(request, 'table/criteria_4_2_1.html', context)


@login_required
def criteria_4_2_1_delete(request):
    Criteria_4_2_1.objects.all().delete()
    return redirect('criteria_4_2_1')


@login_required
def criteria_4_3_1_form(request):
    # Get the existing instance (if any)
    instance = Criteria_4_3_1.objects.first()

    if request.method == 'POST':
        form = CriteriaForm_4_3_1(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('criteria_4_3_1')
        else:
            print(form.errors)
    else:
        form = CriteriaForm_4_3_1(instance=instance)

    return render(request, 'form/criteria_4_3_1.html', {'form': form})


@login_required
def criteria_4_3_1(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='4_3_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="4_3_1")
    context = {
        'criteria_data': criteria_data,
        'has_access': has_access,
        'users': users,
    }
    if not criteria_data:
        context['no_data'] = True
    return render(request, 'table/criteria_4_3_1.html', context)


@login_required
def criteria_4_3_1_delete(request):
    Criteria_4_3_1.objects.all().delete()
    return redirect('criteria_4_3_1')


@login_required
def criteria_4_4_1(request):
    has_access = UserCriteriaLink.objects.filter(user=request.user, criteria='4_4_1').exists()
    users = UserCriteriaLink.objects.filter(criteria="4_4_1")

    print(users)
    records_by_year = {}
    for year in [2021, 2022, 2023, 2024, 2025]:
        records = Criteria_4_4_1.objects.filter(year=year)
        total_amount = records.aggregate(total_amount=Sum('amount_in_lakhs'))['total_amount'] or 0
        records_by_year[f'{year}_records'] = {
            'records': records,
            'total': total_amount
        }
    context = {
        'has_access': has_access,
        'users': users,
        **records_by_year,
    }
    return render(request, 'table/criteria_4_4_1.html', context=context)


@login_required
def criteria_4_4_1_form(request):
    if request.method == 'POST':
        form = CriteriaForm_4_4_1(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data Added Successfully')
            print("ADSf")
            return redirect('criteria_4_4_1')
        else:
            messages.error(request, 'Enter valid Data')
    form = CriteriaForm_4_4_1()
    context = {
        'form': form,
    }
    return render(request, 'form/criteria_4_4_1.html', context=context)


@login_required
def upload_excel_4_4_1(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Extracting values from Excel row
                    year = row[0]
                    head_of_expenditure = row[1]
                    item_of_expenditure = row[2]
                    amount_in_lakhs = row[3]

                    # Creating the Criteria_4_4_1 object
                    Criteria_4_4_1.objects.create(
                        year=year,
                        head_of_expenditure=head_of_expenditure,
                        item_of_expenditure=item_of_expenditure,
                        amount_in_lakhs=amount_in_lakhs
                    )

            messages.success(request, 'Excel file uploaded and processed successfully.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
        return redirect('criteria_4_4_1')
    messages.error(request, 'Please upload a valid Excel file.')
    return redirect('criteria_4_4_1')


@login_required
def export_excel_4_4_1(request):
    data = Criteria_4_4_1.objects.all().order_by('year')

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Criteria 4.4.1"

    headers = [
        'Head of expenditure (for example, Repair and maintenance)',
        'Item of expenditure (for example, AMC for Lab equipment and computers, garden maintenance etc.)',
        'Amount (INR in Lakhs)'
    ]

    column_width = 21.89
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = column_width

    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='top')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    bold_font = Font(bold=True)

    yellow_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

    title_cell = worksheet.cell(row=1, column=1, value='4.4.1 Percentage expenditure incurred on maintenance of physical facilities and academic support facilities excluding salary component, during the last five years')

    start_row = 3

    for year in Criteria_4_4_1.YEAR_CHOICES:
        year_value, year_label = year

        worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
        year_cell = worksheet.cell(row=start_row, column=1, value=f"Year: {year_label}")
        year_cell.alignment = center_alignment
        year_cell.font = bold_font
        year_cell.fill = yellow_fill

        for row in worksheet.iter_rows(min_row=start_row, max_row=start_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        start_row += 1

        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col_num, value=header)
            cell.alignment = wrap_alignment
            cell.border = thin_border
            cell.font = bold_font
        start_row += 1

        year_data = data.filter(year=year_value)
        for entry in year_data:
            cells = [
                worksheet.cell(row=start_row, column=1, value=entry.head_of_expenditure),
                worksheet.cell(row=start_row, column=2, value=entry.item_of_expenditure),
                worksheet.cell(row=start_row, column=3, value=entry.amount_in_lakhs),
            ]
            for cell in cells:
                cell.alignment = wrap_alignment
                cell.border = thin_border
                cell.font = bold_font
            start_row += 1
        worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
        summary_cell = worksheet.cell(row=start_row, column=1, value="Total")
        summary_amount_cell = worksheet.cell(row=start_row, column=3,
                                             value=f"=SUM(C{start_row - len(year_data)}:C{start_row - 1})")

        summary_cell.alignment = wrap_alignment
        summary_cell.border = thin_border
        summary_cell.font = bold_font
        summary_amount_cell.alignment = wrap_alignment
        summary_amount_cell.border = thin_border
        summary_amount_cell.font = bold_font
        start_row += 2

    virtual_workbook = BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response = HttpResponse(
        virtual_workbook,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=4_4_1_criteria_data.xlsx'
    return response

